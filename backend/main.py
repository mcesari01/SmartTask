"""Main application file for the SmartTask backend."""

from typing import List, Optional
import csv
import io
from datetime import datetime, timedelta
from auth import refresh_access_token_with_refresh_token, save_google_tokens_for_user
from jose import JWTError, jwt

from fastapi import Depends, FastAPI, HTTPException, status, Body,Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, RedirectResponse
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy import case, desc, text, inspect
from sqlalchemy.orm import Session
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from auth import (
    create_access_token,
    get_current_user,
    get_db,
    get_password_hash,
    verify_password,
    verify_google_id_token_and_get_email, 
    refresh_access_token_with_refresh_token, 
    save_google_tokens_for_user,
    exchange_code_for_tokens
)
from database.database import Base, engine
from models import Task as TaskModel, User
from schemas.schemas import (
    TaskCreate,
    TaskRead,
    Token,
    UserCreate,
    UserRead,
    TaskCompletedUpdate,
    GoogleSaveToken,
    CalendarEventCreate,
)
from pydantic import BaseModel

app = FastAPI()

# Configura CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],  # Frontend dev origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Crea tabelle
Base.metadata.create_all(bind=engine)

# Lightweight migration: ensure new columns exist in SQLite when upgrading
try:
    inspector = inspect(engine)
    # tasks table: add completed (already present in your file), add google_event_id if missing
    try:
        task_columns = [c["name"] for c in inspector.get_columns("tasks")]
        with engine.connect() as conn:
            if "completed" not in task_columns:
                conn.execute(text("ALTER TABLE tasks ADD COLUMN completed BOOLEAN DEFAULT 0"))
                conn.commit()
            if "google_event_id" not in task_columns:
                conn.execute(text("ALTER TABLE tasks ADD COLUMN google_event_id TEXT"))
                conn.commit()
            if "all_day" not in task_columns:
                # default 0 (False)
                conn.execute(text("ALTER TABLE tasks ADD COLUMN all_day BOOLEAN DEFAULT 0"))
                conn.commit()
    except Exception:
        pass

    # users table: add google token fields if missing
    try:
        user_columns = [c["name"] for c in inspector.get_columns("users")]
        with engine.connect() as conn:
            if "google_access_token" not in user_columns:
                conn.execute(text("ALTER TABLE users ADD COLUMN google_access_token TEXT"))
                conn.commit()
            if "google_refresh_token" not in user_columns:
                conn.execute(text("ALTER TABLE users ADD COLUMN google_refresh_token TEXT"))
                conn.commit()
            if "google_token_expiry" not in user_columns:
                conn.execute(text("ALTER TABLE users ADD COLUMN google_token_expiry DATETIME"))
                conn.commit()
    except Exception:
        pass
except Exception:
    # Best-effort; avoid crashing the app if introspection fails
    pass


@app.post("/register", response_model=UserRead)
def register(user: UserCreate, db: Session = Depends(get_db)):
    """Registers a new user."""
    db_user = db.query(User).filter(User.email == user.email).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    hashed_password = get_password_hash(user.password)
    db_user = User(email=user.email, hashed_password=hashed_password)
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user


@app.post("/login", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    """Authenticates a user and returns an access token."""
    user = db.query(User).filter(User.email == form_data.username).first()
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token = create_access_token(data={"sub": user.email})
    return {"access_token": access_token, "token_type": "bearer"}


class GoogleAuthPayload(BaseModel):
    credential: str


@app.post("/auth/google", response_model=Token)
def google_auth(payload: GoogleAuthPayload, db: Session = Depends(get_db)):
    """Login/Register via Google ID token. Creates user if not exists, then issues JWT."""
    email = verify_google_id_token_and_get_email(payload.credential)

    user = db.query(User).filter(User.email == email).first()
    if not user:
        user = User(email=email, hashed_password="", auth_provider="google")
        db.add(user)
        db.commit()
        db.refresh(user)

    access_token = create_access_token(data={"sub": email})
    return {"access_token": access_token, "token_type": "bearer"}


@app.get("/me")
def get_me(current_user: User = Depends(get_current_user)):
    """
    Return minimal user info for frontend usage.
    NOTE: returns google_access_token (string) if present to match frontend checks.
    If you prefer not to expose the token, change to return only a boolean.
    """
    return {
        "id": current_user.id,
        "email": current_user.email,
        "google_access_token": current_user.google_access_token,
        "google_connected": bool(current_user.google_access_token),
    }

@app.get("/auth/google-calendar/connect")
def google_calendar_connect():
    """Generate Google OAuth URL for calendar connection."""
    from urllib.parse import urlencode
    import os

    client_id = os.getenv("GOOGLE_CLIENT_ID", "39094537919-fpqfsor5dc2mrgbacotk4tlt1mc8j19u.apps.googleusercontent.com")
    redirect_uri = os.getenv("GOOGLE_REDIRECT_URI", "http://localhost:8000/auth/google-calendar/callback")

    params = {
        "client_id": client_id,
        "redirect_uri": redirect_uri,
        "scope": "https://www.googleapis.com/auth/calendar.events",
        "response_type": "code",
        "access_type": "offline",
        "prompt": "consent"
    }

    oauth_url = "https://accounts.google.com/o/oauth2/v2/auth?" + urlencode(params)
    print("🌐 Generated Google OAuth URL:", oauth_url)

    return {
        "oauth_url": oauth_url,
        "client_id": client_id,
        "redirect_uri": redirect_uri
    }

@app.get("/auth/google-calendar/callback")
def google_calendar_callback(
    code: str,
    state: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Callback OAuth che scambia il codice con i token e li salva per l'utente loggato."""
    from auth import exchange_code_for_tokens, save_google_tokens_for_user, SECRET_KEY, ALGORITHM
    from models import User

    # Decodifica il JWT passato come state per capire chi è l'utente
    if not state:
        raise HTTPException(status_code=400, detail="Missing state (user token)")

    try:
        from urllib.parse import unquote
        payload = jwt.decode(unquote(state), SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
        if not email:
            raise HTTPException(status_code=401, detail="Invalid state payload")
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Invalid JWT in state: {str(e)}")

    # Recupera l'utente dal database
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Scambia il code con i token
    token_data = exchange_code_for_tokens(code)

    access_token = token_data.get("access_token")
    refresh_token = token_data.get("refresh_token")
    expires_in = token_data.get("expires_in")

    if not access_token:
        raise HTTPException(status_code=400, detail="No access_token returned from Google")

    save_google_tokens_for_user(db, user, access_token, refresh_token, expires_in)

    return RedirectResponse(url="http://localhost:5173/") 


@app.post("/google-auth")
def save_google_token(
    payload: GoogleSaveToken,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Receives an access_token (and optionally refresh_token/expires_in) from frontend (OAuth flow)
    and stores it for the current user.
    """
    if not payload.access_token:
        raise HTTPException(status_code=400, detail="access_token required")

    current_user.google_access_token = payload.access_token
    if payload.refresh_token:
        current_user.google_refresh_token = payload.refresh_token
    if payload.expires_in:
        try:
            expires_in_seconds = int(payload.expires_in)
            current_user.google_token_expiry = datetime.utcnow() + timedelta(seconds=expires_in_seconds)
        except Exception:
            # ignore parse errors
            pass

    db.add(current_user)
    db.commit()
    db.refresh(current_user)
    return {"detail": "google token saved"}


@app.get("/google-auth/status")
def get_google_auth_status(current_user: User = Depends(get_current_user)):
    """Get the current Google authentication status for the user."""
    from auth import is_valid_refresh_token
    from datetime import datetime
    
    status = {
        "google_connected": bool(current_user.google_access_token),
        "has_refresh_token": bool(current_user.google_refresh_token),
        "refresh_token_valid_format": False,
        "token_expiry": current_user.google_token_expiry.isoformat() if current_user.google_token_expiry else None,
        "token_expired": False,
        "recommendations": []
    }
    
    if current_user.google_refresh_token:
        status["refresh_token_valid_format"] = is_valid_refresh_token(current_user.google_refresh_token)
        
        if not status["refresh_token_valid_format"]:
            status["recommendations"].append("Refresh token appears to be invalid or a placeholder. Please reconnect your Google account.")
    
    if current_user.google_token_expiry:
        status["token_expired"] = current_user.google_token_expiry < datetime.utcnow()
        if status["token_expired"]:
            status["recommendations"].append("Access token has expired. A refresh will be attempted automatically on next API call.")
    
    if not current_user.google_access_token:
        status["recommendations"].append("No Google access token found. Please connect your Google account through the OAuth flow.")
    
    return status


@app.post("/google-auth/refresh")
def manually_refresh_google_token(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Manually refresh the Google access token using the stored refresh token."""
    if not current_user.google_refresh_token:
        raise HTTPException(status_code=400, detail="No refresh token available. Please reconnect your Google account.")
    
    try:
        refreshed = refresh_access_token_with_refresh_token(current_user.google_refresh_token)
        new_access = refreshed.get("access_token")
        expires_in = refreshed.get("expires_in", 3600)
        
        save_google_tokens_for_user(db, current_user, new_access, expires_in=expires_in)
        
        return {
            "detail": "Token refreshed successfully",
            "expires_in": expires_in,
            "new_expiry": current_user.google_token_expiry.isoformat() if current_user.google_token_expiry else None
        }
    except HTTPException as e:
        raise e
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Unexpected error during token refresh: {str(e)}")


@app.post("/google-calendar/events")
def create_google_event(
    event: CalendarEventCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Crea o aggiorna un evento nel calendario Google dell'utente."""
    token_to_use = current_user.google_access_token
    if not token_to_use:
        raise HTTPException(status_code=400, detail="Google account non collegato")

    headers = {"Authorization": f"Bearer {token_to_use}", "Content-Type": "application/json"}

    # Normalizza start/end: se sono dict, usali direttamente; se sono Pydantic, estrai con .model_dump()
    def to_dict_safe(value):
        if hasattr(value, "model_dump"):
            return value.model_dump()
        elif isinstance(value, dict):
            return value
        else:
            raise HTTPException(status_code=400, detail="Invalid event date/time format")

    start_data = to_dict_safe(event.start)
    end_data = to_dict_safe(event.end)

    body = {
        "summary": event.summary,
        "description": event.description or "",
        "start": start_data,
        "end": end_data,
    }


    task = None
    event_id = None
    if event.task_id:
        task = db.query(TaskModel).filter(
            TaskModel.id == event.task_id, TaskModel.user_id == current_user.id
        ).first()
        if not task:
            raise HTTPException(status_code=404, detail="Task not found or not authorized")
        event_id = task.google_event_id

    if event_id:
        # Update existing event (PATCH for partial update)
        url = f"https://www.googleapis.com/calendar/v3/calendars/primary/events/{event_id}"
        resp = requests.patch(url, json=body, headers=headers)
    else:
        # Create new event (POST)
        url = "https://www.googleapis.com/calendar/v3/calendars/primary/events"
        resp = requests.post(url, json=body, headers=headers)

    if resp.status_code == 401 and current_user.google_refresh_token:
        try:
            refreshed = refresh_access_token_with_refresh_token(current_user.google_refresh_token)
            new_access = refreshed.get("access_token")
            expires_in = refreshed.get("expires_in", 3600)
            save_google_tokens_for_user(db, current_user, new_access, expires_in=expires_in)

            # Retry with new token
            headers["Authorization"] = f"Bearer {new_access}"
            if event_id:
                resp = requests.patch(url, json=body, headers=headers)
            else:
                resp = requests.post(url, json=body, headers=headers)
        except HTTPException as e:
            raise e
        except Exception as e:
            raise HTTPException(status_code=401, detail=f"Google token refresh failed: {str(e)}")

    if not resp.ok:
        raise HTTPException(status_code=resp.status_code, detail=f"Google Calendar API error: {resp.text}")

    event_data = resp.json()
    new_event_id = event_data.get('id')

    # If creating new and task provided, save the event ID to the task
    if task and not event_id and new_event_id:
        task.google_event_id = new_event_id
        # If the event was an all-day event (start contains 'date') mark task as all_day
        try:
            if isinstance(start_data, dict) and start_data.get('date'):
                task.all_day = True
                # set deadline to that date at 00:00:00 (local naive datetime)
                from datetime import datetime
                task.deadline = datetime.fromisoformat(start_data.get('date'))
        except Exception:
            pass
        db.commit()
        db.refresh(task)

    return event_data



@app.get("/tasks", response_model=List[TaskRead])
def get_tasks(
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = "asc",
    completed: Optional[str] = None,  # values: 'true' | 'false' | None (all)
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retrieves tasks for the current user, with optional sorting and completion filter.

    sort_by: 'insertion' | 'deadline' | 'priority' | None
    sort_order: 'asc' | 'desc'
    completed: 'true' | 'false' | None
    """
    query = db.query(TaskModel).filter(TaskModel.user_id == current_user.id)

    # Filter by completion if requested
    if completed is not None:
        if completed.lower() == "true":
            query = query.filter(TaskModel.completed.is_(True))
        elif completed.lower() == "false":
            query = query.filter(TaskModel.completed.is_(False))

    # Sorting logic
    if sort_by == "priority":
        priority_order = case(
            (TaskModel.priority == "High", 3),
            (TaskModel.priority == "Medium", 2),
            (TaskModel.priority == "Low", 1),
            else_=0,
        )
        query = query.order_by(priority_order.desc() if sort_order == "desc" else priority_order.asc())
    elif sort_by == "deadline":
        # Interpreting "desc" as: items with closer deadlines should come first.
        query = query.order_by(TaskModel.deadline.asc() if sort_order == "desc" else TaskModel.deadline.desc())
    else:
        query = query.order_by(TaskModel.id.asc() if sort_order == "asc" else TaskModel.id.desc())

    return query.all()


@app.post("/tasks", response_model=TaskRead)
def create_task(
    task: TaskCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Creates a new task for the current user."""
    payload = task.model_dump()
    # ensure all_day default
    if 'all_day' not in payload or payload.get('all_day') is None:
        payload['all_day'] = False
    db_task = TaskModel(**payload, user_id=current_user.id)
    db.add(db_task)
    db.commit()
    db.refresh(db_task)
    return db_task


@app.get("/tasks/{task_id}", response_model=TaskRead)
def get_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retrieves a specific task by ID for the current user."""
    task = db.query(TaskModel).filter(
        TaskModel.id == task_id, TaskModel.user_id == current_user.id
    ).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found or not authorized")
    return task


@app.put("/tasks/{task_id}", response_model=TaskRead)
def update_task(
    task_id: int,
    updated_task: TaskCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Updates an existing task by ID for the current user."""
    task = db.query(TaskModel).filter(
        TaskModel.id == task_id, TaskModel.user_id == current_user.id
    ).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found or not authorized")
    for key, value in updated_task.model_dump().items():
        # preserve existing values if None provided
        if value is None:
            continue
        setattr(task, key, value)
    db.commit()
    db.refresh(task)
    return task


@app.patch("/tasks/{task_id}/completed", response_model=TaskRead)
def set_task_completed(
    task_id: int,
    payload: TaskCompletedUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Toggle or set task completion state for the current user's task."""
    task = db.query(TaskModel).filter(
        TaskModel.id == task_id, TaskModel.user_id == current_user.id
    ).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found or not authorized")
    task.completed = payload.completed
    db.commit()
    db.refresh(task)
    return task


@app.delete("/tasks/{task_id}")
def delete_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Deletes a task by ID for the current user."""
    task = db.query(TaskModel).filter(
        TaskModel.id == task_id, TaskModel.user_id == current_user.id
    ).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found or not authorized")
    db.delete(task)
    db.commit()
    return {"detail": "Task deleted"}


@app.get("/tasks/export/csv")
def export_tasks_csv(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exports all tasks for the current user as CSV."""
    tasks = db.query(TaskModel).filter(TaskModel.user_id == current_user.id).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(['ID', 'Title', 'Description', 'Deadline', 'Priority', 'Completed'])
    
    # Write tasks
    for task in tasks:
        writer.writerow([
            task.id,
            task.title,
            task.description or '',
            task.deadline.strftime('%Y-%m-%d %H:%M:%S') if task.deadline else '',
            task.priority,
            'Yes' if task.completed else 'No'
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename=tasks.csv'}
    )


@app.get("/tasks/export/excel")
def export_tasks_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exports all tasks for the current user as Excel."""
    tasks = db.query(TaskModel).filter(TaskModel.user_id == current_user.id).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"
    
    # Style for header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write header
    headers = ['ID', 'Title', 'Description', 'Deadline', 'Priority', 'Completed']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write tasks
    for row, task in enumerate(tasks, 2):
        ws.cell(row=row, column=1, value=task.id)
        ws.cell(row=row, column=2, value=task.title)
        ws.cell(row=row, column=3, value=task.description or '')
        ws.cell(row=row, column=4, value=task.deadline.strftime('%Y-%m-%d %H:%M:%S') if task.deadline else '')
        ws.cell(row=row, column=5, value=task.priority)
        ws.cell(row=row, column=6, value='Yes' if task.completed else 'No')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=tasks.xlsx'}
    )


@app.get("/tasks/export/pdf")
def export_tasks_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Exports all tasks for the current user as PDF."""
    tasks = db.query(TaskModel).filter(TaskModel.user_id == current_user.id).all()
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title = Paragraph("SmartTask - Export Tasks", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Table data
    data = [['ID', 'Title', 'Description', 'Deadline', 'Priority', 'Completed']]
    
    for i, task in enumerate(tasks, start=1):
        data.append([
            str(i),
            task.title,
            task.description or '',
            task.deadline.strftime('%Y-%m-%d %H:%M') if task.deadline else '',
            task.priority,
            'Yes' if task.completed else 'No'
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/pdf',
        headers={'Content-Disposition': 'attachment; filename=tasks.pdf'}
    )
